
import argparse
import os
import re
import shutil
import unicodedata
from pathlib import Path
from copy import copy

import openpyxl
from pypdf import PdfReader


def normalize_name(name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def sanitize_filename(text: str) -> str:
    text = normalize_name(text)
    text = text.replace(" ", "_")
    text = re.sub(r"[^A-Z0-9_.-]", "", text)
    return text


def parse_brl_number(text: str):
    if text is None or text == "":
        return None
    cleaned = str(text).replace("R$", "").replace(" ", "").strip()
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return round(float(cleaned), 2)
    except Exception:
        return None


def load_reference_map(reference_xlsx: str, sheet_name: str = "GERAL"):
    wb = openpyxl.load_workbook(reference_xlsx, data_only=True)
    ws = wb[sheet_name]

    header_row = None
    socios_col = None
    total_col = None

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            key = normalize_name(val)
            if key == "SOCIOS":
                header_row = row
                socios_col = col
            if key.startswith("TOTAL"):
                total_col = col
        if header_row and total_col:
            break

    if not header_row or not socios_col or not total_col:
        raise ValueError("Não foi possível localizar as colunas 'Sócios' e 'TOTAL' na aba GERAL.")

    result = {}
    for row in range(header_row + 1, ws.max_row + 1):
        nome = ws.cell(row, socios_col).value
        total = ws.cell(row, total_col).value
        if nome is None:
            continue

        nome_txt = str(nome).strip()
        nome_norm = normalize_name(nome_txt)
        if not nome_norm or nome_norm == "TOTAL":
            continue

        if isinstance(total, (int, float)):
            total_val = round(float(total), 2)
        else:
            total_val = parse_brl_number(total)

        result[nome_norm] = {
            "nome_original": nome_txt,
            "valor_pago": total_val,
        }

    return result


def extract_pdf_data(pdf_path: str):
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    nome_match = re.search(r"NOME:\s*(.+)", text)
    cpf_match = re.search(r"CPF:\s*([0-9.\-]+)", text)
    valor_match = re.search(
        r"4\.\s*Rendimentos isentos e não tributáveis\s*\(Distribuição de lucros\)\s*R\$\s*([\d\.\,]+)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    nome = nome_match.group(1).strip() if nome_match else None
    cpf = cpf_match.group(1).strip() if cpf_match else None
    valor = parse_brl_number(valor_match.group(1)) if valor_match else None

    return {
        "arquivo": os.path.basename(pdf_path),
        "caminho": pdf_path,
        "nome_pdf": nome,
        "nome_norm": normalize_name(nome),
        "cpf": cpf,
        "valor_informado": valor,
    }


def find_template_columns(ws):
    header_row = None
    cols = {}

    mapping = {
        "NOME": "nome",
        "CPF": "cpf",
        "VALOR PAGO": "valor_pago",
        "INFORMADO": "informado",
        "DIFERENCA": "diferenca",
        "DIFERENÇA": "diferenca",
        "OBSERVACAO": "observacao",
        "OBSERVAÇÃO": "observacao",
    }

    for row in range(1, min(ws.max_row, 20) + 1):
        current = {}
        for col in range(1, ws.max_column + 1):
            key = normalize_name(ws.cell(row, col).value)
            if key in mapping:
                current[mapping[key]] = col
        if {"nome", "cpf", "valor_pago", "informado"}.issubset(current.keys()):
            header_row = row
            cols = current
            break

    if not header_row:
        raise ValueError("Não foi possível localizar os cabeçalhos Nome / CPF / Valor pago / informado.")

    return header_row, cols


def clone_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def ensure_observation_column(ws, header_row, cols):
    if "observacao" in cols:
        return cols["observacao"]

    obs_col = max(cols.values()) + 1
    dst = ws.cell(header_row, obs_col)
    dst.value = "observação"

    source_col = cols.get("diferenca", cols["informado"])
    src = ws.cell(header_row, source_col)
    clone_cell_style(src, dst)

    src_letter = openpyxl.utils.get_column_letter(source_col)
    dst_letter = openpyxl.utils.get_column_letter(obs_col)
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width or 25

    cols["observacao"] = obs_col
    return obs_col


def get_formula_template(ws, row, col):
    value = ws.cell(row, col).value
    return value if isinstance(value, str) and value.startswith("=") else None


def shift_formula_to_row(formula: str, source_row: int, target_row: int) -> str:
    return re.sub(rf"([A-Z]+){source_row}\b", rf"\g<1>"+str(target_row), formula)


def clear_data_rows(ws, start_row):
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def write_output_workbook(template_xlsx, output_xlsx, rows):
    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb[wb.sheetnames[0]]

    header_row, cols = find_template_columns(ws)
    ensure_observation_column(ws, header_row, cols)

    start_row = header_row + 1
    style_row = start_row

    formula_template = None
    if "diferenca" in cols:
        formula_template = get_formula_template(ws, style_row, cols["diferenca"])

    clear_data_rows(ws, start_row)

    # garante pelo menos algumas larguras úteis
    if ws.column_dimensions["A"].width is None:
        ws.column_dimensions["A"].width = 40
    if ws.column_dimensions["B"].width is None:
        ws.column_dimensions["B"].width = 16
    if ws.column_dimensions["C"].width is None:
        ws.column_dimensions["C"].width = 14
    if ws.column_dimensions["D"].width is None:
        ws.column_dimensions["D"].width = 14

    for idx, item in enumerate(rows, start=start_row):
        for col in range(1, ws.max_column + 1):
            clone_cell_style(ws.cell(style_row, col), ws.cell(idx, col))

        ws.cell(idx, cols["nome"]).value = item.get("nome")
        ws.cell(idx, cols["cpf"]).value = item.get("cpf")
        ws.cell(idx, cols["valor_pago"]).value = item.get("valor_pago")
        ws.cell(idx, cols["informado"]).value = item.get("informado")
        ws.cell(idx, cols["observacao"]).value = item.get("observacao")

        if "diferenca" in cols and formula_template:
            ws.cell(idx, cols["diferenca"]).value = shift_formula_to_row(formula_template, style_row, idx)

        ws.cell(idx, cols["cpf"]).number_format = "@"
        ws.cell(idx, cols["valor_pago"]).number_format = 'R$ #,##0.00'
        ws.cell(idx, cols["informado"]).number_format = 'R$ #,##0.00'
        if "diferenca" in cols:
            ws.cell(idx, cols["diferenca"]).number_format = 'R$ #,##0.00'

    wb.save(output_xlsx)


def process(reference_xlsx, pdf_dir, template_xlsx, output_dir, output_xlsx_name="consolidado_preenchido.xlsx"):
    ref_map = load_reference_map(reference_xlsx)

    pdf_dir = Path(pdf_dir)
    output_dir = Path(output_dir)
    renamed_dir = output_dir / "renomeado"
    renamed_dir.mkdir(parents=True, exist_ok=True)

    rows = []
    report_lines = []
    matched_names = set()

    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    for pdf_file in pdf_files:
        info = extract_pdf_data(str(pdf_file))
        ref = ref_map.get(info["nome_norm"])

        if ref:
            nome_saida = ref["nome_original"]
            valor_pago = ref["valor_pago"]
            observacao = ""
            matched_names.add(info["nome_norm"])
        else:
            nome_saida = info["nome_pdf"] or pdf_file.stem
            valor_pago = None
            observacao = "PDF não encontrado na planilha GERAL"

        novo_nome = f'{sanitize_filename(nome_saida)}_{pdf_file.name}'
        shutil.copy2(pdf_file, renamed_dir / novo_nome)

        rows.append({
            "nome": nome_saida,
            "cpf": info["cpf"],
            "valor_pago": valor_pago,
            "informado": info["valor_informado"],
            "observacao": observacao,
        })

        if ref:
            diff = None
            if valor_pago is not None and info["valor_informado"] is not None:
                diff = round(valor_pago - info["valor_informado"], 2)
            report_lines.append(
                f'OK/CONFERIR: {nome_saida} | Valor pago: {valor_pago} | Informado PDF: {info["valor_informado"]} | Diferença: {diff}'
            )
        else:
            report_lines.append(
                f'PDF SEM CADASTRO NA GERAL: {nome_saida} | CPF: {info["cpf"]} | Informado PDF: {info["valor_informado"]}'
            )

    rows.sort(key=lambda x: normalize_name(x["nome"]))

    missing_in_pdf = sorted(
        [v["nome_original"] for k, v in ref_map.items() if k not in matched_names],
        key=normalize_name
    )

    report_lines.append("")
    report_lines.append("NOMES DA PLANILHA GERAL SEM PDF CORRESPONDENTE:")
    if missing_in_pdf:
        for name in missing_in_pdf:
            report_lines.append(f"- {name}")
    else:
        report_lines.append("- Nenhum")

    output_xlsx = output_dir / output_xlsx_name
    write_output_workbook(template_xlsx, str(output_xlsx), rows)

    report_path = output_dir / "relatorio_conferencia.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    return output_xlsx, report_path, renamed_dir


def main():
    parser = argparse.ArgumentParser(
        description="Confere informes em PDF, renomeia arquivos e preenche planilha consolidada."
    )
    parser.add_argument("referencia_xlsx", help="Planilha de referência com aba GERAL e colunas Sócios/TOTAL")
    parser.add_argument("pdf_dir", help="Pasta com os PDFs")
    parser.add_argument("template_xlsx", help="Planilha modelo do consolidado")
    parser.add_argument("--output-dir", default="saida_processamento", help="Pasta de saída")
    parser.add_argument("--output-xlsx-name", default="consolidado_preenchido.xlsx", help="Nome da planilha gerada")
    args = parser.parse_args()

    output_xlsx, report_path, renamed_dir = process(
        args.referencia_xlsx,
        args.pdf_dir,
        args.template_xlsx,
        args.output_dir,
        args.output_xlsx_name,
    )

    print(f"Planilha gerada: {output_xlsx}")
    print(f"Relatório: {report_path}")
    print(f"PDFs renomeados: {renamed_dir}")


if __name__ == "__main__":
    main()
