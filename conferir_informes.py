import re
import sys
import unicodedata
import shutil
from pathlib import Path
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader

VALOR_TOLERANCIA = Decimal('0.01')


@dataclass
class ExcelSocio:
    nome_original: str
    nome_norm: str
    total: Decimal
    linha: int


@dataclass
class PdfInfo:
    arquivo: Path
    nome_pdf: Optional[str]
    nome_norm: Optional[str]
    valor_pdf: Optional[Decimal]
    texto: str
    erro: Optional[str] = None


@dataclass
class Resultado:
    pdf: PdfInfo
    socio_excel: Optional[ExcelSocio]
    status: str
    observacao: str = ''


def normalizar_nome(texto: str) -> str:
    texto = texto.strip()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper()
    texto = re.sub(r'[^A-Z0-9 ]+', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto


def sanitizar_nome_arquivo(texto: str) -> str:
    texto = normalizar_nome(texto).replace(' ', '_')
    texto = re.sub(r'_+', '_', texto).strip('_')
    return texto[:140] if len(texto) > 140 else texto


def decimal_brasileiro(valor) -> Optional[Decimal]:
    if valor is None or valor == '':
        return None
    if isinstance(valor, (int, float)):
        return Decimal(str(valor)).quantize(Decimal('0.01'))
    s = str(valor).strip()
    s = s.replace('R$', '').strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return Decimal(s).quantize(Decimal('0.01'))
    except InvalidOperation:
        return None


def formatar_decimal_br(valor: Optional[Decimal]) -> str:
    if valor is None:
        return ''
    s = f'{valor:,.2f}'
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def carregar_socios_excel(caminho_excel: Path, aba: str = 'GERAL') -> Dict[str, ExcelSocio]:
    wb = load_workbook(caminho_excel, data_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(f'Aba {aba!r} não encontrada no Excel.')
    ws = wb[aba]

    col_nome = None
    col_total = None
    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                txt = normalizar_nome(v)
                if txt == 'SOCIOS':
                    col_nome = c
                    header_row = r
                if txt.startswith('TOTAL'):
                    col_total = c
                    header_row = r if header_row is None else header_row
        if col_nome and col_total:
            break

    if not col_nome or not col_total:
        raise ValueError('Não encontrei as colunas Sócios e TOTAL na aba GERAL.')

    socios = {}
    for r in range((header_row or 1) + 1, ws.max_row + 1):
        nome = ws.cell(r, col_nome).value
        total = ws.cell(r, col_total).value
        if nome is None:
            continue
        nome_str = str(nome).strip()
        if not nome_str:
            continue
        nome_norm = normalizar_nome(nome_str)
        if nome_norm in {'TOTAL', 'TOTAIS'}:
            continue
        total_dec = decimal_brasileiro(total)
        total_dec = total_dec if total_dec is not None else Decimal('0.00')
        socios[nome_norm] = ExcelSocio(nome_original=nome_str, nome_norm=nome_norm, total=total_dec, linha=r)
    return socios


def extrair_texto_pdf(caminho_pdf: Path) -> str:
    reader = PdfReader(str(caminho_pdf))
    textos = []
    for page in reader.pages:
        textos.append(page.extract_text() or '')
    return '\n'.join(textos)


def extrair_nome_e_valor(texto: str) -> Tuple[Optional[str], Optional[Decimal]]:
    texto_limpo = texto.replace('\xa0', ' ')
    nome = None
    valor = None

    m_nome = re.search(r'NOME\s*:\s*(.*?)\s*CPF\s*:', texto_limpo, re.IGNORECASE | re.DOTALL)
    if m_nome:
        nome = re.sub(r'\s+', ' ', m_nome.group(1)).strip()

    padroes_valor = [
        r'4\.\s*Rendimentos isentos e n[aã]o tribut[aá]veis\s*\(Distribui[cç][aã]o de lucros\)\s*R\$\s*([\d\.,]+)',
        r'4\.\s*Rendimentos isentos e n[aã]o tribut[aá]veis\s*\(Distribui[cç][aã]o de lucros\)\s*([\d\.,]+)',
    ]
    for padrao in padroes_valor:
        m_val = re.search(padrao, texto_limpo, re.IGNORECASE | re.DOTALL)
        if m_val:
            valor = decimal_brasileiro(m_val.group(1))
            break

    return nome, valor


def ler_pdf_info(caminho_pdf: Path) -> PdfInfo:
    try:
        texto = extrair_texto_pdf(caminho_pdf)
        nome, valor = extrair_nome_e_valor(texto)
        return PdfInfo(
            arquivo=caminho_pdf,
            nome_pdf=nome,
            nome_norm=normalizar_nome(nome) if nome else None,
            valor_pdf=valor,
            texto=texto,
            erro=None,
        )
    except Exception as e:
        return PdfInfo(caminho_pdf, None, None, None, '', str(e))


def comparar_pdfs_excel(socios: Dict[str, ExcelSocio], pdf_infos: List[PdfInfo]) -> Tuple[List[Resultado], List[ExcelSocio]]:
    resultados = []
    encontrados = set()

    for pdf in pdf_infos:
        if pdf.erro:
            resultados.append(Resultado(pdf, None, 'ERRO_PDF', pdf.erro))
            continue
        if not pdf.nome_norm:
            resultados.append(Resultado(pdf, None, 'NAO_EXTRAIDO', 'Nome não encontrado no PDF.'))
            continue

        socio = socios.get(pdf.nome_norm)
        if socio is None:
            resultados.append(Resultado(pdf, None, 'NAO_ENCONTRADO_EXCEL', 'Nome do PDF não encontrado na planilha.'))
            continue

        encontrados.add(pdf.nome_norm)
        if pdf.valor_pdf is None:
            resultados.append(Resultado(pdf, socio, 'VALOR_NAO_EXTRAIDO', 'Valor do campo 4 não encontrado no PDF.'))
            continue

        diferenca = abs(pdf.valor_pdf - socio.total)
        if diferenca <= VALOR_TOLERANCIA:
            resultados.append(Resultado(pdf, socio, 'OK', 'Nome e valor conferem.'))
        else:
            obs = f'Excel: {formatar_decimal_br(socio.total)} | PDF: {formatar_decimal_br(pdf.valor_pdf)} | Dif.: {formatar_decimal_br(diferenca)}'
            resultados.append(Resultado(pdf, socio, 'VALOR_DIVERGENTE', obs))

    faltantes = [s for nome_norm, s in socios.items() if nome_norm not in encontrados]
    return resultados, faltantes


def copiar_renomeados(resultados: List[Resultado], pasta_saida: Path) -> List[Tuple[Path, Path]]:
    pasta_saida.mkdir(parents=True, exist_ok=True)
    copiados = []
    for r in resultados:
        if r.socio_excel is None:
            continue
        socio_safe = sanitizar_nome_arquivo(r.socio_excel.nome_original)
        novo_nome = f'{socio_safe}_{r.pdf.arquivo.name}'
        destino = pasta_saida / novo_nome
        contador = 2
        while destino.exists():
            destino = pasta_saida / f'{socio_safe}_{contador}_{r.pdf.arquivo.name}'
            contador += 1
        shutil.copy2(r.pdf.arquivo, destino)
        copiados.append((r.pdf.arquivo, destino))
    return copiados


def gerar_relatorio_txt(caminho: Path, resultados: List[Resultado], faltantes: List[ExcelSocio], copiados: List[Tuple[Path, Path]]):
    linhas = []
    linhas.append('RELATORIO DE CONFERENCIA DE INFORMES')
    linhas.append('=' * 60)
    linhas.append('')
    linhas.append(f'Total de PDFs lidos: {len(resultados)}')
    linhas.append(f'Arquivos renomeados/copied: {len(copiados)}')
    linhas.append(f'PDFs com status OK: {sum(1 for r in resultados if r.status == "OK")}')
    linhas.append(f'PDFs com divergencia de valor: {sum(1 for r in resultados if r.status == "VALOR_DIVERGENTE")}')
    linhas.append(f'Nomes da planilha sem PDF correspondente: {len(faltantes)}')
    linhas.append('')

    linhas.append('DETALHAMENTO DOS PDFS')
    linhas.append('-' * 60)
    for r in resultados:
        linhas.append(f'Arquivo: {r.pdf.arquivo.name}')
        linhas.append(f'Status: {r.status}')
        if r.pdf.nome_pdf:
            linhas.append(f'Nome no PDF: {r.pdf.nome_pdf}')
        if r.socio_excel:
            linhas.append(f'Nome na planilha: {r.socio_excel.nome_original}')
            linhas.append(f'Total planilha: {formatar_decimal_br(r.socio_excel.total)}')
        if r.pdf.valor_pdf is not None:
            linhas.append(f'Valor no PDF: {formatar_decimal_br(r.pdf.valor_pdf)}')
        if r.observacao:
            linhas.append(f'Observação: {r.observacao}')
        linhas.append('')

    linhas.append('NOMES DA PLANILHA SEM PDF')
    linhas.append('-' * 60)
    if faltantes:
        for socio in faltantes:
            linhas.append(f'{socio.nome_original} | Total: {formatar_decimal_br(socio.total)} | Linha Excel: {socio.linha}')
    else:
        linhas.append('Nenhum nome faltante.')

    linhas.append('')
    linhas.append('ARQUIVOS RENOMEADOS')
    linhas.append('-' * 60)
    if copiados:
        for origem, destino in copiados:
            linhas.append(f'{origem.name} -> {destino.name}')
    else:
        linhas.append('Nenhum arquivo copiado.')

    caminho.write_text('\n'.join(linhas), encoding='utf-8')


def processar(caminho_excel: Path, pasta_pdfs: Path, pasta_saida: Optional[Path] = None):
    if pasta_saida is None:
        pasta_saida = pasta_pdfs / 'renomeado'

    socios = carregar_socios_excel(caminho_excel)
    pdfs = sorted(pasta_pdfs.glob('*.pdf'))
    if not pdfs:
        raise ValueError(f'Nenhum PDF encontrado em: {pasta_pdfs}')

    pdf_infos = [ler_pdf_info(p) for p in pdfs]
    resultados, faltantes = comparar_pdfs_excel(socios, pdf_infos)
    copiados = copiar_renomeados(resultados, pasta_saida)
    gerar_relatorio_txt(pasta_saida / 'relatorio_conferencia.txt', resultados, faltantes, copiados)

    print(f'Concluído. PDFs processados: {len(resultados)}')
    print(f'Pasta de saída: {pasta_saida}')
    print(f'Relatório: {pasta_saida / "relatorio_conferencia.txt"}')


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python conferir_informes.py <arquivo_excel.xlsx> <pasta_pdfs> [pasta_saida]')
        sys.exit(1)

    excel = Path(sys.argv[1])
    pasta = Path(sys.argv[2])
    saida = Path(sys.argv[3]) if len(sys.argv) >= 4 else None
    processar(excel, pasta, saida)
